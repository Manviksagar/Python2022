import openpyxl
wb=openpyxl.load_workbook("TestData.xlsx")
sh=wb["Data"]
row=sh.max_row
col=sh.max_column
print(row)
print(col)
for i in range(1,row+1):
    for j in range(1,col + 1):
        print(sh.cell(i,j).value)
