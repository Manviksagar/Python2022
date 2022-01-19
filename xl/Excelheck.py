import openpyxl
wb=openpyxl.load_workbook("TestData.xlsx")
name=wb.sheetnames
print(name)

data=wb['Data']['A2'].value
print(data)
#anotherway...

sh=wb['Data']
d=sh['A2'].value
print(d)

#by cell values

dd=sh.cell(2,2).value
print(dd)
#by dict
ss=sh.cell(row=2,column=2).value
print(ss)